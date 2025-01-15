import requests
from bs4 import BeautifulSoup
import pandas as pd
import numpy as np
import os
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time
import psycopg2
#os.chdir('C:/Users/trann/Documents/IUT/sem 3/SAE/BDD')
os.chdir('U:/Documents/Sem3/SAE/BDD')
url = "https://fr.wikipedia.org/wiki/Friends"
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

## Fonctions:
def get_element(url, tag, class_name):
    page = requests.get(url)
    soup = BeautifulSoup(page.text, 'html.parser')
    element = soup.find(tag, class_ = class_name)
    if element:
        return element
    else:
        return "Erreur"

def extract_birthday_from_url(row):
    url_act = "https://fr.wikipedia.org" + row['URL_Act']
    url_per = "https://fr.wikipedia.org" + row['URL_Per']
    class_name = 'nowrap bday'
    tag = 'time'
    bday_act = get_element(url_act, tag, class_name).text
    bday_per = get_element(url_per, tag, class_name).text
    return pd.Series([bday_act, bday_per], index=['bday_act', 'bday_per'])

def extract_birthday_from_url2(url_wiki):
    class_name = 'nowrap bday'
    tag = 'time'
    bday_act = get_element(url_wiki, tag, class_name)
    if bday_act == "Erreur":
        class_name = 'nowrap date-lien bday'
        bday_act = get_element(url_wiki, tag, class_name)
    if bday_act == "Erreur":
        class_name = 'nowrap date-lien'
        bday_act = get_element(url_wiki, tag, class_name)
    if bday_act == "Erreur":
        bday_act = get_element(url_wiki, 'span', 'bday')
    if bday_act != "Erreur":
        return bday_act.text
    else:
        return None

def get_birthday_from_name_wiki(name):
    #chrome_options = Options()
    #chrome_options.add_argument("--headless")
    #driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://www.wikipedia.org/")
    search_box = WebDriverWait(driver,1).until(
    EC.element_to_be_clickable((By.NAME, "search"))
    )
    search_box.send_keys(name)
    search_box.send_keys(Keys.RETURN)
    url_wiki = driver.current_url
    try:
        birthday = extract_birthday_from_url2(url_wiki)
    except Exception as e:
        birthday = None
    #driver.quit()
    return birthday
def get_birthday_from_name_duckduckgo(name):
    #chrome_options = Options()
    #chrome_options.add_argument("--headless")
    #driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://www.duckduckgo.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(name + " actor wikipedia")
    search_box.send_keys(Keys.RETURN)
    time.sleep(2)
    try:
        first_result = driver.find_element(By.XPATH, "(//a[contains(@href, 'wikipedia.org')])[1]")
        if first_result:
            driver.execute_script("arguments[0].click();", first_result)
            time.sleep(5)
            url_wiki = driver.current_url
            birthday = extract_birthday_from_url2(url_wiki)
        else:
            birthday = None
    except Exception as e:
        birthday = None
    #driver.quit()
    return birthday

data = []
def collect_data_from_page(html,group):

    soup = BeautifulSoup(html, 'html.parser')
    paragraphs = soup.find_all('p')
    for p in paragraphs:
        b_tags = p.find_all('b')
        if len(b_tags) >= 2:
            col1 = b_tags[0].get_text(strip=True)
            col2 = b_tags[1].get_text(strip=True)

            episode_links = []
            for a in p.find_all('a', href=True):
                if 'episode' in a['href']:
                    href_value = a['href'].split('=')[-1]
                    episode_links.append(href_value)

            if "Cliquez" not in col1 and "Cliquez" not in col2:
                data.append([col1, col2, episode_links,group])

## Acteurs principaux
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')
table = soup.find('table', class_='wikitable centre')
table = get_element(url,'table','wikitable centre')
liste_act = []
liste_URL_act = []
liste_per = []
liste_URL_per = []
if table:
    rows = table.find_all('tr')
    for row in rows:
        cols = row.find_all('td')
        if len(cols) >= 2:
            col1 = cols[0].get_text()
            a_tag = cols[0].find('a')
            if a_tag:
                link_col1 = a_tag.get('href')
            col2 = cols[1].get_text()
            a_tag2 = cols[1].find('a')
            if a_tag2:
                link_col2 = a_tag2.get('href')
            liste_act.append(col1)
            liste_URL_act.append(link_col1)
            liste_per.append(col2)
            liste_URL_per.append(link_col2)
else:
    print("Erreur")
df_acteur = pd.DataFrame({
    'Act': liste_act,
    'URL_Act': liste_URL_act,
    'Per': liste_per,
    'URL_Per': liste_URL_per
})


# Appliquer la fonction pour extraire les dates de naissance
df_acteur[['bday_act', 'bday_per']] = df_acteur.apply(extract_birthday_from_url, axis=1)
df_acteur['stt_invite'] = 0
# Afficher le DataFrame final avec les dates de naissance
print(df_acteur)
#Vérifier
print(df_acteur.iloc[1])

## Acteurs secondaires

url = "https://www.fanfr.com/invites/friendsgeneration2.php?importance=Celebrite&actioninv=Rechercher"

response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')

paragraphs = soup.find_all('p')

data = []
for p in paragraphs:
    b_tags = p.find_all('b')
    if len(b_tags) >= 2:
        col1 = b_tags[0].get_text(strip=True)
        col2 = b_tags[1].get_text(strip=True)

        episode_links = []
        for a in p.find_all('a', href=True):
            if 'episode' in a['href']:
                href_value = a['href'].split('=')[-1]
                episode_links.append(href_value)

        if "Cliquez" not in col1 and "Cliquez" not in col2:
            data.append([col1, col2, episode_links])
df = pd.DataFrame(data)
df.columns = ['personnage', 'acteur','episode']
df.to_csv("friends_data.csv", index=False, sep=";", encoding='utf-8')
## Récupérer des liens des invités
url = "https://www.fanfr.com/invites/"
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')
list_link = []
links = soup.find_all('a',href=lambda href: href and 'friendsgeneration2.php?importance=' in href)
for link in links:
    result_link = "https://www.fanfr.com/invites/"+link['href']
    list_link.append(result_link)

# Parcourir les liens pour récupérer les données

data = []

chrome_options = Options()
chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)

try:
    for url in list_link:
        group = url.split('importance=')[-1].split('&action')[0]
        print(f"On est à: {url}")

        driver.get(url)
        time.sleep(3)

        collect_data_from_page(driver.page_source,group)

        while True:
            try:
                next_button = driver.find_element(By.LINK_TEXT, "les invités suivants")
                next_button.click()
                time.sleep(3)
                collect_data_from_page(driver.page_source,group)
            except:
                print("Done and next:")
                break

finally:
    driver.quit()
df = pd.DataFrame(data)
df.columns = ['personnage', 'acteur','episode','group']
df.to_csv("friends_data.csv", index=False, sep=";", encoding='utf-8')
## Chercher la date de naissance avec wiki

data_acteur = pd.read_csv("friends_data.csv", sep = ";")
data_acteur["date_naissance"] = None

chrome_options = Options() chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)

# Boucle pour traiter chaque nom
for i in range(len(data_acteur)):
    if i % 50 == 0:
        print("On traite la ligne", i+1)
    name = data_acteur.iloc[i,1] if 'non crédité' in name:
        print(name)
    else:
        birthday = get_birthday_from_name_wiki(name)

        if birthday:
            data_acteur.iloc[i,4] = birthday
        else:
            print(name)
# Fermer le navigateur
driver.quit() data_acteur.to_csv("friends_data_updated1201.csv",
index=False, sep=";", encoding='utf-8')
## Chercher la date de naissance p2
data_acteur = pd.read_csv("friends_data_updated1201.csv", sep = ";")
chrome_options = Options()
#chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)
comp = 0
# Boucle pour traiter chaque nom
for i in range(len(data_acteur)):
    if i % 50 == 0:
        print("On traite la ligne", i+1)
    bday = data_acteur.iloc[i,4]
    name = data_acteur.iloc[i,1]
    if type(bday) == float:

        if "non crédité" in name:
            print(name)
        else:

            birthday = get_birthday_from_name_duckduckgo(name)
            if birthday:
                data_acteur.iloc[i,4] = birthday
                comp += 1

            else:
                print(name)
print(comp)
driver.quit()
data_acteur.to_csv("friends_data_updated1201_2.csv", index=False, sep=";", encoding='utf-8')
## Nettoyer les dates de naissance de acteur
data_acteur = pd.read_csv("friends_data_updated1201_2.csv", sep = ";")
data_acteur["date_naissance_nettoyee"] = None
ensemble_mois= {
    "janvier": "01", "février": "02", "mars": "03", "avril": "04", "mai": "05",
    "juin": "06", "juillet": "07", "août": "08", "septembre": "09", "octobre": "10",
    "novembre": "11", "décembre": "12"
}
for i in range(len(data_acteur)):
    bday = data_acteur.iloc[i,4]
    if type(bday) == float:
        data_acteur.iloc[i,5] = None
    elif len(bday) == 10:
        data_acteur.iloc[i,5] = bday
    else:
        bday_list = bday.split()
        ensemble_mois[bday_list[1]]
        if bday_list[0] == '1er':
            bday = "01/" + ensemble_mois[bday_list[1]] + "/" + bday_list[2]

        else:
            bday = bday_list[0] + "/" + ensemble_mois[bday_list[1]] + "/" + bday_list[2]
        data_acteur.iloc[i,5] = bday

data_acteur.to_csv("friends_data_updated1201_2.csv", index=False, sep=";", encoding='utf-8')

## Chercher la date de naissance de personnage
data_acteur = pd.read_csv("friends_data_updated1201_2.csv", sep = ";")
data_acteur["date_naissance_personnage"] = None #data_acteur.columns[10]

chrome_options = Options()
#chrome_options.add_argument("--headless")
driver = webdriver.Chrome(options=chrome_options)
# Boucle pour traiter chaque nom
for i in range(117):
#for i in range(6):
    if i % 50 == 0:
        print("On traite la ligne", i+1)
    name = data_acteur.iloc[i,1]
    if 'non crédité' in name:
        print(name)
    else:
        birthday = get_birthday_from_name_duckduckgo(name)

        if birthday:
            data_acteur.iloc[i,10] = birthday
        else:
            print(name)
# Fermer le navigateur
driver.quit()
data_acteur.to_csv("friends_data_updated1201_2.csv",index=False, sep=";", encoding='utf-8')

## Nettoyer les dates de naissance de personnage
data_acteur = pd.read_csv("friends_data_updated1201_2.csv", sep = ";")
data_acteur["date_naissance_nettoyee"] = None
ensemble_mois= {
    "janvier": "01", "février": "02", "mars": "03", "avril": "04", "mai": "05",
    "juin": "06", "juillet": "07", "août": "08", "septembre": "09", "octobre": "10",
    "novembre": "11", "décembre": "12"
}
for i in range(len(data_acteur)):
    bday = data_acteur.iloc[i,4]
    if type(bday) == float:
        data_acteur.iloc[i,5] = None
    elif len(bday) == 10:
        data_acteur.iloc[i,5] = bday
    else:
        bday_list = bday.split()
        ensemble_mois[bday_list[1]]
        if bday_list[0] == '1er':
            bday = "01/" + ensemble_mois[bday_list[1]] + "/" + bday_list[2]

        else:
            bday = bday_list[0] + "/" + ensemble_mois[bday_list[1]] + "/" + bday_list[2]
        data_acteur.iloc[i,5] = bday

data_acteur.to_csv("friends_data_updated1201_2.csv", index=False, sep=";", encoding='utf-8')
## Nettoyer les noms et prénoms
data_acteur = pd.read_csv("friends_data_updated1201_2.csv", sep = ";")

data_acteur['nom_personnage'] = None #data_acteur.columns[6]
data_acteur['prenom_personnage'] = None #data_acteur.columns[7]
data_acteur['nom_acteur'] = None #data_acteur.columns[8]
data_acteur['prenom_acteur'] = None #data_acteur.columns[9]
for i in range(len(data_acteur)):
    nom_prenoms_personnage = data_acteur.iloc[i,1]
    list_noms_p = nom_prenoms_personnage.split()
    if len(list_noms_p) == 1:
        nom_p = nom_prenoms_personnage
    else:
        nom_p = list_noms_p[-1]
        index_nom_p = - (len(nom_p) + 1)
        prenom_p = nom_prenoms_personnage[:index_nom_p]

    nom_prenoms_acteur = data_acteur.iloc[i,2]
    list_noms_a = nom_prenoms_acteur.split()
    if len(list_noms_a) == 1:
        nom_a = nom_prenoms_acteur
    else:
        nom_a = list_noms_a[-1]
        index_nom_a = - (len(nom_a) + 1)
        prenom_a = nom_prenoms_acteur[:index_nom_a]

    data_acteur.iloc[i,6] = nom_p
    data_acteur.iloc[i,7] = prenom_p
    data_acteur.iloc[i,8] = nom_a
    data_acteur.iloc[i,9] = prenom_a
data_acteur.to_csv("friends_data_updated1201_2.csv", index=False, sep=";", encoding='utf-8')

## Créer le table contenir
data_acteur = pd.read_csv("friends_data_updated1201_2.csv", sep = ";")
episodes = pd.read_excel('episode.xlsx')
columns = ['id_episode', 'id_personnage']
contenir = pd.DataFrame(columns=columns)
for i in range(len(data_acteur)):
    list_ep = data_acteur.iloc[i,3]
    if list_ep == "*":
        for id_ep in range(len(episodes)):
            contenir.loc[len(contenir)] = [episodes.iloc[id_ep,0],data_acteur.iloc[i,0]]

    else:
        list_ep = list_ep.strip("[]").replace("'","")
        list_ep = list_ep.replace(" ","").split(',')
        for ep in list_ep:
            contenir.loc[len(contenir)] = [ep,data_acteur.iloc[i,0]]
contenir.to_csv("contenir.csv", index=False, sep=";", encoding='utf-8')

## Alimenter la base de donnée
# Connexion à la base PostgreSQL
ma_connection = psycopg2.connect(
    database="postgres",
    user="postgres",
    host="localhost",
    password="nhung",
    port="5434"
)
## Connexion à la base PostgreSQL à l'IUT
ma_connection = psycopg2.connect(
    database="2025_SAE_Nhung_Jo_Artur",
    user="admindbetu",
    host='10.11.159.10',
    password="admindbetu",
    port="5432"
)
## TABLE ACTEUR ----> à ajouter la date de naissance dans db
mon_curseur = ma_connection.cursor()
# Charger le fichier Excel
workbook = load_workbook(filename="acteur.xlsx")
sheet = workbook.active  # Par défaut, on utilise la première feuille

# Lire les données de la feuille Excel et créer une liste de tuples
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (entêtes)
    data.append(row)  # Chaque ligne devient un tuple (id_acteur, statut_invite, date_naissance_acteur, nom_acteur, prenom_acteur)

# Insérer les données dans la table
query = """
INSERT INTO acteur (id_acteur, statut_invite, date_naissance_acteur, nom_acteur, prenom_acteur)
VALUES (%s, %s, %s, %s, %s)
"""
mon_curseur.executemany(query, data)  # Utilisation de la liste de tuples

# Valider les changements dans la base
ma_connection.commit()
print("Données insérées avec succès !")

mon_curseur.close()
ma_connection.close()
## TABLE PERSONNAGE ---> à changer la longeur de prénom dans db
mon_curseur = ma_connection.cursor()
# Charger le fichier Excel
workbook = load_workbook(filename="personnage.xlsx")
sheet = workbook.active
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(row)

query = """
INSERT INTO personnage (id_personnage, nom_personnage, prenom_personnage, date_naissance_personnage)
VALUES (%s, %s, %s, %s)
"""
mon_curseur.executemany(query, data)  # Utilisation de la liste de tuples

# Valider les changements dans la base
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
## TABLE SAIS0N---> à changer la longeur de resume_saison dans db et type de annee_saison dans fichier excel
# Charger le fichier Excel

mon_curseur = ma_connection.cursor()
workbook = load_workbook(filename="saison.xlsx")
sheet = workbook.active
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(row)
query = """
INSERT INTO saison (id_saison, nom_saison, annee_saison, resume_saison)
VALUES (%s, %s, %s, %s)
"""
mon_curseur.executemany(query, data)
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
## TABLE EPISODE---> à changer la longeur de resume_episode et nom_episode dans db

mon_curseur = ma_connection.cursor()
workbook = load_workbook(filename="episode.xlsx")
sheet = workbook.active  # Par défaut, on utilise la première feuille

# Lire les données de la feuille Excel et créer une liste de tuples
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorer la première ligne (entêtes)
    data.append(row)  # Chaque ligne devient un tuple (id_acteur, statut_invite, date_naissance_acteur, nom_acteur, prenom_acteur)

# Insérer les données dans la table
query = """
INSERT INTO episode (id_episode, num_episode, id_saison, nom_episode, resume_episode)
VALUES (%s, %s, %s, %s, %s)
"""
mon_curseur.executemany(query, data)  # Utilisation de la liste de tuples

# Valider les changements dans la base
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
## TABLE CONTENIR

mon_curseur = ma_connection.cursor()
data = []

with open("contenir.csv" ,mode="r", encoding="utf-8") as csv_file:
    csv_reader = csv.DictReader(csv_file,  delimiter=";")  # Lire en tant que dictionnaires
    for row in csv_reader:
        data.append((
            row['id_episode'],
            row['id_personnage']
        ))

# Insérer les données dans la table
query = """
INSERT INTO contenir (id_episode, id_personnage)
VALUES (%s, %s)
"""
mon_curseur.executemany(query, data)  # Utilisation de la liste de tuples

# Valider les changements dans la base
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
## TABLE INTERPRETER

mon_curseur = ma_connection.cursor()
workbook = load_workbook(filename="interpreter.xlsx")
sheet = workbook.active
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(row)
query = """
INSERT INTO interpreter (id_personnage, id_acteur)
VALUES (%s, %s)
"""
mon_curseur.executemany(query, data)
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
## TABLE SITE WEB #sheet evaluer, effacer les formules de note_saison: 4.5*2

workbook = load_workbook(filename="site_noter_evaluer.xlsx")

sheet_site_web = workbook['site_web']
sheet_noter = workbook['noter']
sheet_evaluer = workbook['evaluer']

data_site_web = []
data_noter = []
data_evaluer = []

for row in sheet_site_web.iter_rows(min_row=2, values_only=True):
    data_site_web.append(row)
for row in sheet_noter.iter_rows(min_row=2, values_only=True):
    data_noter.append(row)
for row in sheet_evaluer.iter_rows(min_row=2, values_only=True):
    data_evaluer.append(row)
query_site_web = """
INSERT INTO site_web (id_site, nom_site)
VALUES (%s, %s)
"""
query_noter = """
INSERT INTO noter (id_episode, id_site, note_episode, nb_votant)
VALUES (%s, %s, %s, %s)
"""
query_evaluer = """
INSERT INTO evaluer (id_saison, id_site, note_saison, nb_votes)
VALUES (%s, %s, %s, %s)
"""
mon_curseur = ma_connection.cursor()
#mon_curseur.executemany(query_site_web, data_site_web)
#mon_curseur.executemany(query_noter, data_noter)
mon_curseur.executemany(query_evaluer, data_evaluer)
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
### TABLE RELATION
mon_curseur = ma_connection.cursor()
workbook = load_workbook(filename="relation.xlsx")
sheet = workbook.active
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    data.append(row)
query = """
INSERT INTO relation (id_relation, type_relation)
VALUES (%s, %s)
"""
mon_curseur.executemany(query, data)
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()
### TABLE LIER
mon_curseur = ma_connection.cursor()
data = []
with open("lier.csv" ,mode="r", encoding="utf-8") as csv_file:
    csv_reader = csv.DictReader(csv_file,  delimiter=";")  # Lire en tant que dictionnaires
    for row in csv_reader:
        data.append((
            row['ID_ep'],
            row['ID_personnage'],
            row['ID_personnage_1'],
            row['ID_relation']
        ))

# Insérer les données dans la table
query = """
INSERT INTO lier (id_episode, id_personnage, id_personnage_1, id_relation)
VALUES (%s, %s, %s, %s)
"""
mon_curseur.executemany(query, data)  # Utilisation de la liste de tuples

# Valider les changements dans la base
ma_connection.commit()
print("Données insérées avec succès !")
mon_curseur.close()
ma_connection.close()